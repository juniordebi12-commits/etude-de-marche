from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Survey, Question, Choice, Respondent, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils.text import slugify
from django.http import HttpResponse, FileResponse, JsonResponse
from django.utils import timezone
import json
import qrcode
from io import BytesIO
import base64
from django.urls import reverse
from wordcloud import WordCloud
from collections import Counter
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib import colors
from xml.sax.saxutils import escape
import tempfile
from .decorators import group_required
from django.contrib.auth.decorators import login_required
from django.http import Http404
from functools import wraps
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError


def get_survey_or_404(request, survey_id):
    user = request.user

    if user.is_staff or user.is_superuser:
        return get_object_or_404(Survey, id=survey_id)

    return get_object_or_404(Survey, id=survey_id, owner=user)


def get_public_survey_or_404(survey_id):
    return get_object_or_404(Survey, id=survey_id)


def generate_wordcloud_and_stats(text_list):
    """Retourne image base64 et top words (liste (mot, count))."""
    full_text = " ".join(text_list).lower()

    if not full_text.strip():
        return None, []

    wc = WordCloud(width=600, height=400, background_color="white").generate(
        full_text
    )
    buffer = BytesIO()
    wc.to_image().save(buffer, format="PNG")
    img_b64 = base64.b64encode(buffer.getvalue()).decode()

    words = [word for word in full_text.split() if len(word) > 3]
    freq = Counter(words).most_common(10)

    return img_b64, freq


def generate_qr_for_survey(request, survey):
    """Retourne data URI PNG du QR code pointant vers la page take_survey."""
    base_url = request.build_absolute_uri("/")[:-1]
    survey_id = survey.id if hasattr(survey, "id") else int(survey)
    qr_url = base_url + reverse("take_survey", args=[survey_id])

    qr = qrcode.make(qr_url)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    return f"data:image/png;base64,{qr_base64}"


@login_required
def survey_list(request):
    user = request.user

    if user.is_staff or user.is_superuser:
        surveys = Survey.objects.all()
    else:
        surveys = Survey.objects.filter(owner=user)

    return render(request, "survey/survey_list.html", {"surveys": surveys})


@login_required
def survey_create(request):
    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        description = request.POST.get("description", "").strip()

        survey = Survey.objects.create(
            title=title,
            description=description,
            owner=request.user,
        )

        question_texts = request.POST.getlist("question_text[]")
        question_types = request.POST.getlist("question_types[]")

        for i, text in enumerate(question_texts):
            if not text or not text.strip():
                continue

            q_type = question_types[i] if i < len(question_types) else "text"

            question = Question.objects.create(
                survey=survey,
                text=text.strip(),
                question_type=q_type,
            )

            if q_type in ["single", "multiple"]:
                choices = request.POST.getlist(f"choices_{i}[]")

                for choice_text in choices:
                    if choice_text and choice_text.strip():
                        Choice.objects.create(
                            question=question,
                            text=choice_text.strip(),
                        )

        return redirect("survey_list")

    return render(request, "survey/survey_create.html")


def survey_summary(request, survey_id):
    survey = get_survey_or_404(request, survey_id)
    respondents_qs = Respondent.objects.filter(survey=survey)

    chart_data = []

    for question in survey.questions.all():
        entry = {
            "text": question.text,
            "type": question.question_type,
        }

        if question.question_type in ["single", "multiple"]:
            labels, values = [], []

            for choice in question.choices.all():
                labels.append(choice.text)

                count = Response.objects.filter(
                    question=question,
                    respondent__in=respondents_qs,
                    selected_choices=choice,
                ).count()

                values.append(count)

            entry["labels"] = labels
            entry["data"] = values

        else:
            answers = Response.objects.filter(
                question=question,
                respondent__in=respondents_qs,
            ).values_list("answer_text", flat=True)

            entry["answers"] = [
                answer for answer in answers if answer and str(answer).strip()
            ]

        chart_data.append(entry)

    context = {
        "survey": survey,
        "chart_data": json.dumps(chart_data, ensure_ascii=False),
        "respondents": respondents_qs,
        "qr_code": generate_qr_for_survey(request, survey),
    }

    text_answers = [
        response.answer_text.strip()
        for response in Response.objects.filter(
            question__survey=survey,
            question__question_type="text",
            respondent__in=respondents_qs,
        )
        if response.answer_text and response.answer_text.strip()
    ]

    if text_answers:
        wc_image, top_words = generate_wordcloud_and_stats(text_answers)

        if wc_image:
            context["wordcloud_img"] = f"data:image/png;base64,{wc_image}"
            context["top_words"] = top_words
    else:
        context["wordcloud_img"] = None
        context["top_words"] = []

    return render(request, "survey/survey_summary.html", context)

def jwt_required(view_func):
    @wraps(view_func)
    def wrapped_view(request, *args, **kwargs):
        # Connexion Django classique : conserve le fonctionnement admin.
        if request.user.is_authenticated:
            return view_func(request, *args, **kwargs)

        try:
            authentication = JWTAuthentication()
            authenticated = authentication.authenticate(request)

            if authenticated is None:
                return JsonResponse(
                    {"detail": "Authentification requise."},
                    status=401,
                )

            request.user, request.auth = authenticated

            return view_func(request, *args, **kwargs)

        except (InvalidToken, TokenError):
            return JsonResponse(
                {"detail": "Jeton invalide ou expiré."},
                status=401,
            )

    return wrapped_view
def get_filtered_export_data(request, survey):
    """Retourne les répondants et réponses limités à la période demandée."""
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")

    respondents_qs = Respondent.objects.filter(survey=survey)

    if date_from:
        respondents_qs = respondents_qs.filter(
            created_at__date__gte=date_from
        )

    if date_to:
        respondents_qs = respondents_qs.filter(
            created_at__date__lte=date_to
        )

    respondents = list(respondents_qs.order_by("-created_at"))

    responses = (
        Response.objects.filter(respondent__in=respondents)
        .select_related("respondent", "question")
        .prefetch_related("selected_choices")
    )

    if date_from and date_to:
        period_label = f"Du {date_from} au {date_to}"
    elif date_from:
        period_label = f"À partir du {date_from}"
    elif date_to:
        period_label = f"Jusqu’au {date_to}"
    else:
        period_label = "Toutes les données"

    return respondents, responses, period_label
@jwt_required
def export_survey_pdf(request, survey_id):
    survey = get_survey_or_404(request, survey_id)

    questions = list(survey.questions.all().prefetch_related("choices"))

    respondents, responses, period_label = get_filtered_export_data(
        request,
        survey,
    )

    answers_by_question = {question.id: [] for question in questions}

    for answer in responses:
        if answer.question.question_type in ("single", "multiple"):
            value = " ; ".join(
                choice.text for choice in answer.selected_choices.all()
            )
        else:
            value = (answer.answer_text or "").strip()

        answers_by_question.setdefault(answer.question_id, []).append(value)

    total_respondents = len(respondents)

    interviewer_counts = Counter(
        respondent.interviewer_name or "Non renseigné"
        for respondent in respondents
    )

    total_interviewers = len(interviewer_counts)
    total_questions = len(questions)
    total_answers = responses.count()

    expected_answers = total_respondents * total_questions

    completion_rate = (
        round((total_answers / expected_answers) * 100, 1)
        if expected_answers > 0
        else 0
    )

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=45,
        bottomMargin=45,
        title=f"Rapport - {survey.title}",
        author="SanaMetrics",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "SanaTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=27,
        textColor=colors.HexColor("#1F4E78"),
        alignment=TA_CENTER,
        spaceAfter=18,
    )

    subtitle_style = ParagraphStyle(
        "SanaSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#5B6573"),
        alignment=TA_CENTER,
        spaceAfter=20,
    )

    section_style = ParagraphStyle(
        "SanaSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#1F4E78"),
        spaceBefore=16,
        spaceAfter=10,
    )

    question_style = ParagraphStyle(
        "SanaQuestion",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#162B3F"),
        spaceBefore=14,
        spaceAfter=8,
    )

    normal_style = ParagraphStyle(
        "SanaNormal",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        alignment=TA_LEFT,
    )

    small_style = ParagraphStyle(
        "SanaSmall",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#5B6573"),
    )

    story = []

    # Page de couverture
    story.append(Spacer(1, 70))
    story.append(Paragraph("SanaMetrics", title_style))
    story.append(Paragraph("Rapport d'analyse d'enquête", subtitle_style))
    story.append(Spacer(1, 30))

    story.append(
        Paragraph(
            escape(survey.title),
            ParagraphStyle(
                "SurveyTitle",
                parent=title_style,
                fontSize=18,
                leading=24,
            ),
        )
    )

    story.append(Spacer(1, 15))

    description = survey.description or "Aucune description fournie."

    story.append(
        Paragraph(
            escape(description),
            ParagraphStyle(
                "CoverDescription",
                parent=normal_style,
                alignment=TA_CENTER,
                fontSize=10,
                leading=15,
            ),
        )
    )

    story.append(Spacer(1, 45))

    cover_data = [
        [
            "Date du rapport",
            timezone.localtime(timezone.now()).strftime("%d/%m/%Y"),
        ],
        ["Nombre de répondants", str(total_respondents)],
        ["Nombre d'enquêteurs", str(total_interviewers)],
        ["Nombre de questions", str(total_questions)],
        ["Taux de complétion", f"{completion_rate} %"],
        ["Période sélectionnée", period_label],
    ]

    cover_table = Table(
        cover_data,
        colWidths=[180, 180],
        hAlign="CENTER",
    )

    cover_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2F3")),
            ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#F5F9FC")),
            ("TOPPADDING", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ])
    )

    story.append(cover_table)
    story.append(PageBreak())

    # Résumé
    story.append(Paragraph("Résumé exécutif", section_style))

    story.append(
        Paragraph(
            (
                f"Cette enquête comprend <b>{total_respondents}</b> répondant(s), "
                f"<b>{total_interviewers}</b> enquêteur(s), "
                f"<b>{total_questions}</b> question(s) et "
                f"<b>{total_answers}</b> réponse(s) enregistrée(s). "
                f"Le taux de complétion estimé est de "
                f"<b>{completion_rate} %</b>."
            ),
            normal_style,
        )
    )

    story.append(Spacer(1, 14))

    summary_data = [
        ["Indicateur", "Valeur"],
        ["Répondants", str(total_respondents)],
        ["Enquêteurs", str(total_interviewers)],
        ["Questions", str(total_questions)],
        ["Réponses enregistrées", str(total_answers)],
        ["Taux de complétion", f"{completion_rate} %"],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[260, 120],
        repeatRows=1,
    )

    summary_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 1), (1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2F3")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                colors.white,
                colors.HexColor("#F5F9FC"),
            ]),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ])
    )

    story.append(summary_table)
    story.append(Spacer(1, 18))

    # Répartition par enquêteur
    story.append(Paragraph("Répartition par enquêteur", section_style))

    interviewer_data = [["Enquêteur", "Répondants collectés"]]

    for interviewer_name, count in interviewer_counts.most_common():
        interviewer_data.append([
            Paragraph(escape(interviewer_name), normal_style),
            str(count),
        ])

    interviewer_table = Table(
        interviewer_data,
        colWidths=[280, 115],
        repeatRows=1,
    )

    interviewer_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 1), (1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2F3")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                colors.white,
                colors.HexColor("#F5F9FC"),
            ]),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ])
    )

    story.append(interviewer_table)

    # Résultats détaillés
    story.append(Paragraph("Résultats détaillés", section_style))

    if not questions:
        story.append(
            Paragraph(
                "Cette enquête ne contient aucune question.",
                normal_style,
            )
        )

    for index, question in enumerate(questions, start=1):
        answers = [
            answer
            for answer in answers_by_question.get(question.id, [])
            if answer
        ]

        answered_count = len(answers)

        story.append(
            Paragraph(
                f"{index}. {escape(question.text)}",
                question_style,
            )
        )

        story.append(
            Paragraph(
                (
                    f"Type : <b>{escape(question.get_question_type_display())}</b>"
                    f" — Réponses obtenues : <b>{answered_count}</b>"
                ),
                small_style,
            )
        )

        story.append(Spacer(1, 6))

        if question.question_type in ("single", "multiple"):
            table_data = [["Réponse", "Effectif", "Pourcentage"]]

            for choice in question.choices.all():
                count = sum(
                    1
                    for answer in answers
                    if choice.text in [
                        item.strip() for item in answer.split(";")
                    ]
                )

                percentage = (
                    round((count / total_respondents) * 100, 1)
                    if total_respondents
                    else 0
                )

                table_data.append([
                    Paragraph(escape(choice.text), normal_style),
                    str(count),
                    f"{percentage} %",
                ])

            if len(table_data) == 1:
                table_data.append(["Aucune modalité définie", "0", "0 %"])

            result_table = Table(
                table_data,
                colWidths=[250, 65, 80],
                repeatRows=1,
            )

            result_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2F3")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                        colors.white,
                        colors.HexColor("#F5F9FC"),
                    ]),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ])
            )

            story.append(result_table)

            if question.question_type == "multiple":
                story.append(
                    Paragraph(
                        "Note : une personne peut sélectionner plusieurs réponses ; le total peut dépasser 100 %.",
                        small_style,
                    )
                )

        elif question.question_type == "number":
            numeric_values = []

            for answer in answers:
                try:
                    numeric_values.append(float(answer.replace(",", ".")))
                except (ValueError, AttributeError):
                    pass

            if numeric_values:
                average = round(
                    sum(numeric_values) / len(numeric_values),
                    2,
                )

                numeric_data = [
                    ["Indicateur", "Valeur"],
                    ["Nombre de réponses", str(len(numeric_values))],
                    ["Moyenne", str(average)],
                    ["Minimum", str(min(numeric_values))],
                    ["Maximum", str(max(numeric_values))],
                ]

                numeric_table = Table(
                    numeric_data,
                    colWidths=[250, 145],
                    repeatRows=1,
                )

                numeric_table.setStyle(
                    TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                        ("ALIGN", (1, 1), (1, -1), "CENTER"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2F3")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                            colors.white,
                            colors.HexColor("#F5F9FC"),
                        ]),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ])
                )

                story.append(numeric_table)

            else:
                story.append(
                    Paragraph(
                        "Aucune réponse numérique exploitable.",
                        normal_style,
                    )
                )

        else:
            if not answers:
                story.append(
                    Paragraph(
                        "Aucune réponse textuelle recueillie.",
                        normal_style,
                    )
                )
            else:
                for answer in answers[:15]:
                    story.append(
                        Paragraph(
                            f"• {escape(answer)}",
                            normal_style,
                        )
                    )
                    story.append(Spacer(1, 3))

                if len(answers) > 15:
                    story.append(
                        Paragraph(
                            (
                                f"{len(answers) - 15} autre(s) réponse(s) "
                                "sont disponibles dans l'export Excel."
                            ),
                            small_style,
                        )
                    )

        story.append(Spacer(1, 12))

    def add_page_number(canvas, doc):
        canvas.saveState()

        canvas.setStrokeColor(colors.HexColor("#D9E2F3"))
        canvas.line(40, 32, A4[0] - 40, 32)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#5B6573"))

        canvas.drawString(40, 20, "SanaMetrics — Rapport d'enquête")

        canvas.drawRightString(
            A4[0] - 40,
            20,
            f"Page {doc.page}",
        )

        canvas.restoreState()

    document.build(
        story,
        onFirstPage=add_page_number,
        onLaterPages=add_page_number,
    )

    safe_title = slugify(survey.title) or f"enquete-{survey.id}"

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
    )

    response["Content-Disposition"] = (
        f'attachment; filename="sanametrics_{safe_title}_rapport.pdf"'
    )

    return response


@jwt_required
def export_survey_excel(request, survey_id):
    survey = get_survey_or_404(request, survey_id)

    questions = list(survey.questions.all().prefetch_related("choices"))

    respondents, responses, period_label = get_filtered_export_data(
        request,
        survey,
    )

    answers_by_respondent = {}
    answers_by_question = {question.id: [] for question in questions}

    for answer in responses:
        if answer.question.question_type in ("single", "multiple"):
            value = " ; ".join(
                choice.text for choice in answer.selected_choices.all()
            )
        else:
            value = (answer.answer_text or "").strip()

        answers_by_respondent.setdefault(
            answer.respondent_id, {}
        ).setdefault(
            answer.question_id, []
        ).append(value)

        answers_by_question.setdefault(answer.question_id, []).append(value)

    workbook = Workbook()

    primary_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="1F4E78")

    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    def excel_datetime(value):
        """Convertit une date Django avec fuseau horaire pour Excel."""
        if not value:
            return None

        if timezone.is_aware(value):
            return timezone.localtime(value).replace(tzinfo=None)

        return value

    def style_header(sheet, row_number=1):
        for cell in sheet[row_number]:
            cell.fill = primary_fill
            cell.font = white_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = thin_border

    def format_table(sheet, header_row=1):
        style_header(sheet, header_row)
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = sheet.dimensions

        for row in sheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)

            max_length = max(
                len(str(cell.value or "")) for cell in column_cells
            )

            sheet.column_dimensions[letter].width = min(
                max(max_length + 2, 14),
                45,
            )

    # -------------------------------------------------------------
    # FEUILLE 1 : SYNTHÈSE
    # -------------------------------------------------------------
    summary_sheet = workbook.active
    summary_sheet.title = "Synthèse"

    total_respondents = len(respondents)
    total_questions = len(questions)
    total_answers = responses.count()
    expected_answers = total_respondents * total_questions

    completion_rate = (
        round((total_answers / expected_answers) * 100, 1)
        if expected_answers > 0
        else 0
    )

    summary_sheet.merge_cells("A1:D1")
    summary_sheet["A1"] = f"Rapport d'enquête : {survey.title}"
    summary_sheet["A1"].font = title_font
    summary_sheet["A1"].alignment = Alignment(horizontal="center")

    summary_sheet["A3"] = "Description"
    summary_sheet["B3"] = survey.description or "Aucune description"

    summary_sheet["A5"] = "Indicateur"
    summary_sheet["B5"] = "Valeur"

    indicators = [
        ("Nombre de répondants", total_respondents),
        ("Nombre de questions", total_questions),
        ("Nombre de réponses", total_answers),
        ("Taux de complétion", f"{completion_rate} %"),
        (
            "Date de génération",
            timezone.localtime(survey.updated_at).strftime("%d/%m/%Y %H:%M"),
        ),
        ("Période sélectionnée", period_label),
    ]

    for row_index, (label, value) in enumerate(indicators, start=6):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=value)

    summary_sheet["A5"].fill = primary_fill
    summary_sheet["B5"].fill = primary_fill
    summary_sheet["A5"].font = white_font
    summary_sheet["B5"].font = white_font

    for row in summary_sheet.iter_rows(
        min_row=5,
        max_row=11,
        min_col=1,
        max_col=2,
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 55

    # -------------------------------------------------------------
    # FEUILLE 2 : RÉPONSES BRUTES
    # -------------------------------------------------------------
    raw_sheet = workbook.create_sheet("Réponses brutes")

    headers = [
        "ID répondant",
        "Participant",
        "Enquêteur",
        "Date de collecte",
        "Statut",
    ] + [question.text for question in questions]

    raw_sheet.append(headers)

    for respondent in respondents:
        respondent_answers = answers_by_respondent.get(respondent.id, {})

        row = [
            respondent.id,
            respondent.participant_name or "Anonyme",
            respondent.interviewer_name or "Non renseigné",
            excel_datetime(respondent.created_at),
            respondent.status,
        ]

        for question in questions:
            values = respondent_answers.get(question.id, [])
            row.append(" ; ".join(value for value in values if value) or "")

        raw_sheet.append(row)

    for row in raw_sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].number_format = "dd/mm/yyyy hh:mm"

    format_table(raw_sheet)

    # -------------------------------------------------------------
    # FEUILLE 3 : STATISTIQUES
    # -------------------------------------------------------------
    stats_sheet = workbook.create_sheet("Statistiques")

    stats_sheet.append([
        "Question",
        "Type",
        "Réponse / Modalité",
        "Effectif",
        "Pourcentage",
    ])

    for question in questions:
        question_answers = [
            answer
            for answer in answers_by_question.get(question.id, [])
            if answer
        ]

        answered_count = len(question_answers)

        if question.question_type in ("single", "multiple"):
            for choice in question.choices.all():
                count = sum(
                    1
                    for answer in question_answers
                    if choice.text in [
                        item.strip() for item in answer.split(";")
                    ]
                )

                percentage = (
                    round((count / total_respondents) * 100, 1)
                    if total_respondents
                    else 0
                )

                stats_sheet.append([
                    question.text,
                    question.get_question_type_display(),
                    choice.text,
                    count,
                    percentage / 100,
                ])

        elif question.question_type == "number":
            numeric_values = []

            for answer in question_answers:
                try:
                    numeric_values.append(float(answer.replace(",", ".")))
                except (ValueError, AttributeError):
                    pass

            if numeric_values:
                statistics = [
                    ("Nombre de réponses", len(numeric_values)),
                    (
                        "Moyenne",
                        round(
                            sum(numeric_values) / len(numeric_values),
                            2,
                        ),
                    ),
                    ("Minimum", min(numeric_values)),
                    ("Maximum", max(numeric_values)),
                ]

                for label, value in statistics:
                    stats_sheet.append([
                        question.text,
                        question.get_question_type_display(),
                        label,
                        value,
                        "",
                    ])
            else:
                stats_sheet.append([
                    question.text,
                    question.get_question_type_display(),
                    "Aucune réponse numérique",
                    0,
                    "",
                ])

        else:
            stats_sheet.append([
                question.text,
                question.get_question_type_display(),
                "Réponses non vides",
                answered_count,
                (
                    answered_count / total_respondents
                    if total_respondents
                    else 0
                ),
            ])

    for row in stats_sheet.iter_rows(min_row=2, min_col=5, max_col=5):
        row[0].number_format = "0.0%"

    format_table(stats_sheet)

    # -------------------------------------------------------------
    # FEUILLE 4 : VERBATIMS
    # -------------------------------------------------------------
    verbatims_sheet = workbook.create_sheet("Verbatims")

    verbatims_sheet.append([
        "Question",
        "Réponse",
        "Participant",
        "Enquêteur",
        "Date de collecte",
    ])

    for answer in responses:
        if (
            answer.question.question_type == "text"
            and answer.answer_text
            and answer.answer_text.strip()
        ):
            verbatims_sheet.append([
                answer.question.text,
                answer.answer_text.strip(),
                answer.respondent.participant_name or "Anonyme",
                answer.respondent.interviewer_name or "Non renseigné",
                excel_datetime(answer.respondent.created_at),
            ])

    for row in verbatims_sheet.iter_rows(min_row=2, min_col=5, max_col=5):
        row[0].number_format = "dd/mm/yyyy hh:mm"

    format_table(verbatims_sheet)

    # -------------------------------------------------------------
    # FEUILLE 5 : DICTIONNAIRE
    # -------------------------------------------------------------
    dictionary_sheet = workbook.create_sheet("Dictionnaire")

    dictionary_sheet.append([
        "Ordre",
        "Question",
        "Type",
        "Choix possibles",
    ])

    for question in questions:
        choices = " ; ".join(
            choice.text for choice in question.choices.all()
        )

        dictionary_sheet.append([
            question.order,
            question.text,
            question.get_question_type_display(),
            choices,
        ])

    format_table(dictionary_sheet)

    safe_title = slugify(survey.title) or f"enquete-{survey.id}"
    filename = f"sanametrics_{safe_title}_export.xlsx"

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    workbook.save(response)

    return response


def delete_all_responses(request, survey_id):
    survey = get_survey_or_404(request, survey_id)

    if request.method == "POST":
        Response.objects.filter(respondent__survey=survey).delete()
        Respondent.objects.filter(survey=survey).delete()

        messages.success(
            request,
            (
                f"Toutes les réponses et tous les répondants pour "
                f"'{survey.title}' ont été supprimés."
            ),
        )

        return redirect("survey_summary", survey_id=survey.id)

    return render(
        request,
        "survey/confirm_delete_responses.html",
        {"survey": survey},
    )


def delete_respondent(request, respondent_id):
    respondent = get_object_or_404(Respondent, id=respondent_id)
    survey = respondent.survey

    if not (
        request.user.is_staff
        or request.user.is_superuser
        or survey.owner == request.user
    ):
        raise Http404

    if request.method == "POST":
        Response.objects.filter(respondent=respondent).delete()
        respondent.delete()

        messages.success(
            request,
            (
                f"Le répondant '{respondent.interviewer_name}' "
                f"et ses réponses ont été supprimés."
            ),
        )

        return redirect("survey_summary", survey_id=survey.id)

    return render(
        request,
        "survey/confirm_delete_respondent.html",
        {
            "respondent": respondent,
            "survey": survey,
        },
    )


def survey_detail(request, survey_id):
    survey = get_survey_or_404(request, survey_id)
    questions = survey.questions.all()

    return render(
        request,
        "survey/detail.html",
        {
            "survey": survey,
            "questions": questions,
        },
    )


def take_survey(request, survey_id):
    survey = get_public_survey_or_404(survey_id)
    questions = survey.questions.all()

    if request.method == "POST":
        interviewer_name = request.POST.get(
            "interviewer_name",
            "",
        ).strip()

        if not interviewer_name:
            interviewer_name = "Réponse en ligne"

        respondent = Respondent.objects.create(
            survey=survey,
            interviewer_name=interviewer_name,
            created_by=survey.owner,
        )

        for question in questions:
            field_name = f"question_{question.id}"

            if question.question_type in ["single", "multiple"]:
                selected_choices = request.POST.getlist(field_name)

                if selected_choices:
                    response_obj = Response.objects.create(
                        respondent=respondent,
                        question=question,
                    )

                    for choice_id in selected_choices:
                        try:
                            choice = Choice.objects.get(id=int(choice_id))
                            response_obj.selected_choices.add(choice)
                        except (Choice.DoesNotExist, ValueError):
                            continue

            else:
                answer_text = request.POST.get(field_name, "").strip()

                if answer_text:
                    Response.objects.create(
                        respondent=respondent,
                        question=question,
                        answer_text=answer_text,
                    )

        if (
            request.headers.get("Content-Type") == "application/json"
            or request.path.startswith("/api/")
        ):
            return JsonResponse({
                "status": "success",
                "respondent_id": respondent.id,
                "interviewer_name": interviewer_name,
            })

        return render(request, "survey/thanks.html", {"survey": survey})

    return render(
        request,
        "survey/take_survey.html",
        {
            "survey": survey,
            "questions": questions,
        },
    )


def survey_edit(request, survey_id):
    survey = get_survey_or_404(request, survey_id)
    has_responses = survey.respondents.exists()

    if request.method == "POST":
        survey.title = request.POST.get("title", "").strip()
        survey.description = request.POST.get("description", "").strip()
        survey.save()

        if has_responses:
            messages.warning(
                request,
                (
                    "Le titre a été mis à jour, mais les questions sont "
                    "verrouillées car des réponses existent déjà."
                ),
            )

            return redirect("survey_detail", survey_id=survey.id)

        for question in survey.questions.all():
            text_key = f"question_text_existing_{question.id}"
            type_key = f"question_type_existing_{question.id}"

            if text_key not in request.POST:
                question.delete()
                continue

            question.text = request.POST.get(
                text_key,
                question.text,
            ).strip()

            question.question_type = request.POST.get(
                type_key,
                question.question_type,
            )

            question.save()

            if question.question_type in ["single", "multiple"]:
                question.choices.all().delete()

                for choice_text in request.POST.getlist(
                    f"choices_existing_{question.id}[]"
                ):
                    if choice_text and choice_text.strip():
                        Choice.objects.create(
                            question=question,
                            text=choice_text.strip(),
                        )

        new_texts = request.POST.getlist("new_question_text[]")
        new_types = request.POST.getlist("new_question_type[]")

        for i, text in enumerate(new_texts):
            if not text or not text.strip():
                continue

            q_type = new_types[i] if i < len(new_types) else "text"

            question = Question.objects.create(
                survey=survey,
                text=text.strip(),
                question_type=q_type,
            )

            if q_type in ["single", "multiple"]:
                for choice_text in request.POST.getlist(
                    f"new_choices_{i}[]"
                ):
                    if choice_text and choice_text.strip():
                        Choice.objects.create(
                            question=question,
                            text=choice_text.strip(),
                        )

        return redirect("survey_detail", survey_id=survey.id)

    return render(
        request,
        "survey/survey_edit.html",
        {
            "survey": survey,
            "has_responses": has_responses,
        },
    )


def survey_delete(request, survey_id):
    survey = get_survey_or_404(request, survey_id)

    if request.method == "POST":
        deleted_title = survey.title
        survey.delete()

        return render(
            request,
            "survey/survey_deleted.html",
            {"deleted_title": deleted_title},
        )

    return render(
        request,
        "survey/survey_confirm_delete.html",
        {"survey": survey},
    )


def survey_thanks(request, survey_id):
    survey = get_survey_or_404(request, survey_id)

    return render(
        request,
        "survey/thanks.html",
        {"survey": survey},
    )